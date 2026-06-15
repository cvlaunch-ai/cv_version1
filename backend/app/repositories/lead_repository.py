import os
import csv
import time
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials

from app.core.config import settings

class LeadRepository:
    @staticmethod
    def save_lead_data(data: dict) -> str:
        """
        Saves lead data. First attempts Google Sheets via credentials.json.
        If credentials.json is missing or an error occurs, falls back to appending
        to a local CSV file: ../database/leads_data.csv.
        """
        # 1. Try Google Sheets
        try:
            scope = [
                "https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/drive"
            ]
            
            # Note: We look for credentials.json in the project root/backend directory
            # as it was originally checked relative to cwd.
            if os.path.exists("credentials.json"):
                creds = ServiceAccountCredentials.from_json_keyfile_name(
                    "credentials.json", scope
                )
                client = gspread.authorize(creds)
                
                # ✅ Open by URL provided by user
                sheet_url = "https://docs.google.com/spreadsheets/d/1_AUHN66wwgHt48MVrc66uzzp8qdA1BTzTh0Gvn8aUpA/edit"
                sheet = client.open_by_url(sheet_url).sheet1
                
                sheet.append_row([
                    data["name"],
                    data["mailid"],
                    data["phone"],
                    data["jobrole"],
                    str(datetime.now())
                ])
                return "[OK] Saved to Google Sheet"
            else:
                raise FileNotFoundError("credentials.json not found")
                
        except Exception as e:
            print(f"Google Sheet Error: {e}")
            
            # 2. Fallback to Local CSV
            os.makedirs(settings.DATABASE_DIR, exist_ok=True)
            csv_file = os.path.join(settings.DATABASE_DIR, "leads_data.csv")
            file_exists = os.path.isfile(csv_file)
            
            timestamp = str(datetime.now())
            
            row_data = [
                data["name"],
                data["mailid"],
                data["phone"],
                data["jobrole"],
                timestamp
            ]

            # Retry mechanism: Try for 15 seconds
            max_retries = 15
            
            for attempt in range(max_retries):
                try:
                    # Try appending to the main file
                    with open(csv_file, "a", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        if not file_exists and attempt == 0: 
                            writer.writerow(["Name", "Email", "Phone", "Job Role", "Timestamp"])
                        writer.writerow(row_data)
                    return "[OK] Saved to CSV"
                
                except PermissionError:
                    # If locked, wait and retry
                    if attempt < max_retries - 1:
                        time.sleep(1) # Wait 1 second
                        continue
                    else:
                        return "[ERROR] FILE LOCKED. Please close 'leads_data.csv' immediately!"
                except Exception as e:
                    return f"[ERROR] Save Failed: {str(e)}"

    @staticmethod
    def save_lead_excel(row_data: list, excel_file: str) -> str:
        """
        Saves lead row data to leads_data.xlsx using openpyxl, with a retry mechanism.
        """
        from openpyxl import load_workbook, Workbook
        max_retries = 15
        for attempt in range(max_retries):
            try:
                # Check if file exists
                if os.path.isfile(excel_file):
                    wb = load_workbook(excel_file)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    # Write header
                    ws.append(["Name", "Email", "Phone", "Job Role", "Timestamp"])
                
                # Append data row
                ws.append(row_data)
                
                # Save workbook
                wb.save(excel_file)
                print(f"Saved to Excel: {row_data}")
                return f"Saved to {excel_file}"
            
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(1)
                    continue
                else:
                    raise PermissionError("File is locked")

