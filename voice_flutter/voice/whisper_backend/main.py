from fastapi import FastAPI, UploadFile, File, WebSocket
from fastapi.middleware.cors import CORSMiddleware
import whisper
import os
import tempfile
import imageio_ffmpeg
import asyncio
import websockets
from websockets.exceptions import ConnectionClosedOK
from deep_translator import GoogleTranslator
from typing import List
import json
from pydantic import BaseModel
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Add current directory to PATH so Whisper finds the local ffmpeg.exe
os.environ["PATH"] = os.getcwd() + os.pathsep + os.environ["PATH"]
print(f"Added to PATH: {os.getcwd()}")

app = FastAPI()

# Allow all origins
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load Whisper model (Base for better performance/memory usage)
print("Loading Whisper model... (this may take a minute on first run)")
model = whisper.load_model("base") 
print("Whisper model loaded successfully!")

@app.post("/transcribe")
async def transcribe_audio(file: UploadFile = File(...)):
    try:
        # Read audio
        audio_bytes = await file.read()
        print(f"Received audio: {len(audio_bytes)} bytes")
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".webm") as temp_audio:
            temp_audio.write(audio_bytes)
            temp_audio_path = temp_audio.name
        
        # Transcribe with PROMPT for better accuracy/pronunciation
        print("Transcribing audio...")
        prompt = "The following is a clear conversation in English, Hindi, or Telugu. Please transcribe it accurately with correct punctuation and spelling."
        
        result = model.transcribe(temp_audio_path, initial_prompt=prompt)
        
        # Cleanup
        os.unlink(temp_audio_path)
        
        text = result["text"].strip()
        print(f"Result: {text}")
        
        return {"text": text}
    
    except Exception as e:
        print(f"Error: {e}")
        return {"text": f"[ERROR] {str(e)}"}

# ================================
# ⚙️ CONFIGURATION & DIRECTORIES
# ================================
import subprocess
import platform

# Use D: drive for all file operations as requested
BASE_CONVERSION_DIR = "D:\\conversions"
DOWNLOADS_DIR = os.path.join(BASE_CONVERSION_DIR, "downloads")

# Ensure directories exist
for d in [BASE_CONVERSION_DIR, DOWNLOADS_DIR]:
    if not os.path.exists(d):
        os.makedirs(d, exist_ok=True)

# Path to LibreOffice (Customize if installed elsewhere)
LIBREOFFICE_PATHS = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    r"D:\Program Files\LibreOffice\program\soffice.exe",
    "soffice"
]

def get_soffice_path():
    for path in LIBREOFFICE_PATHS:
        if path == "soffice" or os.path.exists(path):
            return path
    return None

def convert_with_libreoffice(input_path: str, output_ext: str):
    """Uses LibreOffice Headless to convert documents with 100% exact layout."""
    soffice = get_soffice_path()
    if not soffice:
        print("[WARN] LibreOffice not found. Falling back to internal engine.")
        return None
        
    try:
        # LibreOffice outputs to the same folder as input or a specified outdir
        outdir = os.path.dirname(input_path)
        cmd = [soffice, "--headless", "--convert-to", output_ext, "--outdir", outdir, input_path]
        print(f"Executing Engine: {' '.join(cmd)}")
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0:
            # Result filename is usually original_name.new_ext
            base_name = os.path.splitext(os.path.basename(input_path))[0]
            output_path = os.path.join(outdir, f"{base_name}.{output_ext}")
            
            if os.path.exists(output_path):
                # Save a copy to D Drive Downloads as requested
                copy_path = os.path.join(DOWNLOADS_DIR, f"{base_name}_{int(time.time())}.{output_ext}")
                shutil.copy2(output_path, copy_path)
                print(f"✅ Exact Match Saved to: {copy_path}")
                return output_path
        
        print(f"LibreOffice Error: {result.stderr}")
        return None
    except Exception as e:
        print(f"Engine Error: {e}")
        return None

@app.get("/")
async def root():
    return {"message": "[MIC] Free Whisper API - Running locally!", "status": "ready"}

# ================================
# ✅ NEW IMPORTS (ADD BELOW YOUR CURRENT IMPORTS)
# ================================
import re
import time
import gspread
import csv
import os
import json
from datetime import datetime
from oauth2client.service_account import ServiceAccountCredentials
from pydantic import BaseModel
import google.generativeai as genai
import openai

# ================================
# ✅ API CLIENT CONFIGURATION
# ================================
# 1. Try Google Gemini (Free Tier Available)
gemini_api_key = os.getenv("GEMINI_API_KEY")
if gemini_api_key:
    genai.configure(api_key=gemini_api_key)
    print("[OK] Gemini API Key found.")

# 2. Try OpenAI (Paid)
# User provided specific key for Resume generation
openai_api_key = os.getenv("OPENAI_API_KEY")
if openai_api_key:
    os.environ["OPENAI_API_KEY"] = openai_api_key # Ensure v1.0 client sees it
    try:
        openai.api_key = openai_api_key
        print("[OK] OpenAI API Key found.")
    except NameError:
        print("[WARN] OpenAI package not installed, skipping OpenAI config.")

if not gemini_api_key and not openai_api_key:
    print("[WARN] No AI API Keys found. Using Regex Fallback.")

# ================================
# ✅ FIXED RESUME TEMPLATES & LOGIC
# ================================

RESUME_TEMPLATE = """
{full_name}
| {phone} | {email} | {location} |

Professional Summary
{summary}

Education
{education}

Roles & Responsibilities
{responsibilities}

Projects
{projects}

Technologies & Tools
{skills}
"""

def parse_formatted_resume(text: str):
    """
    100% Literal parser for UI text.
    Everything between headers is captured exactly.
    """
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines: return {}
    
    data = {"name": lines[0], "phone": "", "email": "", "location": ""}
    
    # Extract contact info from line 1
    if len(lines) > 1:
        # Split by pipes but handle multiple pipes or empty parts
        parts = [p.strip() for p in lines[1].split("|") if p.strip()]
        for p in parts:
            if "@" in p: data["email"] = p
            elif any(c.isdigit() for c in p) and len(p.replace(" ","")) > 6: 
                data["phone"] = p
            else: data["location"] = p

    # Mapping for headers (Lowercase for matching)
    header_map = {
        "professional summary": "summary",
        "professional_summary": "summary",
        "summary": "summary",
        "education": "education",
        "roles & responsibilities": "responsibilities",
        "roles and responsibilities": "responsibilities",
        "experience": "responsibilities",
        "work experience": "responsibilities",
        "projects": "projects",
        "technologies & tools": "skills",
        "technical skills": "skills",
        "skills": "skills",
        "certifications": "certifications"
    }

    sections = {k: [] for k in ["summary", "education", "responsibilities", "projects", "skills", "certifications"]}
    current_key = "summary" 
    
    for line in lines[2:]:
        # Clean line for header checking
        clean_l = line.lower().replace(":", "").replace("*","").strip()
        
        # Check if line IS a header
        is_header = False
        if clean_l in header_map:
            current_key = header_map[clean_l]
            is_header = True
            
        if not is_header and current_key:
            sections[current_key].append(line)
            
    # Combine lines back to strings
    for k, v in sections.items():
        data[k] = "\n".join(v).strip()
        
    return data

def parse_structured_resume(text: str):
    """
    Parses a semi-structured text dump that has headers like 'Education', 'Experience'.
    This is common when users copy-paste their resume or speak it section by section.
    """
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    if not lines:
        return {}

    data = {
        "name": "",
        "email": "",
        "phone": "",
        "location": "",
        "summary": "",
        "education": [],
        "projects": [],
        "skills": [],
        "summary": "",
        "education": [],
        "projects": [],
        "skills": [],
        "responsibilities": [],
        "certifications": []
    }

    # 1. Header Extraction (First few lines often contain Name + Contact)
    # Heuristic: Name is usually the first line.
    
    header_end_index = 0
    # Check lines for contact details
    # We scan more lines (up to 15) because some resumes have long headers
    for i, line in enumerate(lines[:15]):
        line_lower = line.lower()
        # Stop at first section header
        if any(h in line_lower for h in ["professional summary", "summary", "education", "experience", "projects", "skills", "certifications"]):
            # Check if it's NOT just a job title in experience (like "Senior UI/UX Designer")
            # Headers are usually short or have specific formatting.
            # But "Experience" is a strong keyword.
            if len(line.split()) < 5: # Headers are usually short
                 header_end_index = i
                 break
        
        # Name Heuristic (First line usually)
        if i == 0:
            name_match = re.search(r"Name:\s*(.*)", line, re.I)
            if name_match:
                data["name"] = name_match.group(1).strip()
            else:
                # If mostly text and short, assume name
                if len(line) < 50 and not any(char.isdigit() for char in line):
                    data["name"] = line
        
        # Contact Heuristics
        email_match = re.search(r"[\w\.-]+@[\w\.-]+\.\w+", line)
        if email_match: data["email"] = email_match.group()
        
        phone_match = re.search(r"(\+?\d[\d -]{8,15})", line)
        if phone_match and len(line) < 40 and "year" not in line_lower: # Avoid "5+ years"
            data["phone"] = phone_match.group(1).strip()
        
        if "location:" in line_lower:
            try:
                 parts = line.split(":", 1)
                 if len(parts) > 1:
                     data["location"] = parts[1].strip()
            except: pass
        elif ("india" in line_lower or "hyderabad" in line_lower) and "@" not in line and "experience" not in line_lower:
             # Heuristic for location line just containing "Hyderabad"
             data["location"] = line.strip()
        
        # Check for Role/Job Title in header
        role_match = re.search(r"(?:Role|Job Title|Position):\s*(.*)", line, re.I)
        if role_match:
             # We can use this to seed the summary if summary is empty
             role = role_match.group(1).strip()
             if not data["summary"]:
                 data["summary"] = f"Aspiring {role} with a strong background in technical skills and a passion for data-driven solutions."

    # 2. Section Parsing using Headers
    current_section = None
    buffer = []
    
    def flush_buffer(section, content_lines):
        if not content_lines: return
        content = "\n".join(content_lines)
        
        if section == "summary":
            data["summary"] = content
        elif section == "education":
            data["education"].append(content)
        elif section == "experience":
            data["responsibilities"].append(content)
        elif section == "projects":
            data["projects"].append(content)
        elif section == "skills":
            # Replace bullets with commas for skills
            cleaned = content.replace('\n', ', ').replace('•', ',')
            data["skills"].append(cleaned)
        elif section == "certifications":
            data["certifications"].append(content)

    section_map = {
        "professional summary": "summary",
        "career summary": "summary",
        "summary": "summary",
        "education": "education",
        "qualification": "education",
        "experience": "experience",
        "work experience": "experience",
        "employment history": "experience",
        "roles & responsibilities": "experience", # Map roles to experience/responsibilities
        "projects": "projects",
        "academic projects": "projects",
        "skills": "skills",
        "technical skills": "skills",
        "certifications": "education",
        "certifications": "certifications",
        "technologies & tools": "skills",
        # Emoji variants
        "🎓 education": "education",
        "🛠️ skills": "skills",
        "📊 projects": "projects",
        "📜 certifications": "certifications",
        "💼 experience": "experience",
        "case projects": "projects"
    }

    for i in range(header_end_index, len(lines)):
        line = lines[i]
        # Clean line for key matching: Remove emojis, punctuation
        clean_key = re.sub(r'[^\w\s]', '', line.lower()).strip()
        
        # Check if line is a header by exact or partial match
        is_header = False
        mapped_section = None
        
        # 1. Exact Match via Map
        if clean_key in section_map:
             is_header = True
             mapped_section = section_map[clean_key]
        
        # 2. Robust Fuzzy Match (If header seems like a title)
        # Check specific keywords if the line is short enough to be a header
        if not is_header and len(clean_key) < 35:
            if "summary" in clean_key: mapped_section, is_header = "summary", True
            elif "education" in clean_key or "qualification" in clean_key: mapped_section, is_header = "education", True
            elif "experience" in clean_key or "work" in clean_key or "employment" in clean_key: mapped_section, is_header = "experience", True
            elif "project" in clean_key: mapped_section, is_header = "projects", True
            elif "skill" in clean_key or "technolog" in clean_key: mapped_section, is_header = "skills", True
            elif "certif" in clean_key: mapped_section, is_header = "education", True
        
        if is_header:
            # Flush previous section
            if current_section:
                flush_buffer(current_section, buffer)
            
            # Start new section
            current_section = mapped_section
            buffer = []
        else:
            if current_section:
                buffer.append(line)
            # If no section yet, check for Summary
            elif not current_section:
                 # Logic for pre-header summaries (like Degree info)
                 # If it's not contact info
                 is_contact = False
                 if "@" in line or "phone" in line.lower() or "portfolio" in line.lower() or "linkedin" in line.lower():
                     is_contact = True
                 if i == 0: is_contact = True # Name line

                 if not is_contact:
                     # Treat as summary
                     current_section = "summary"
                     buffer.append(line)
    
    # Flush last section
    if current_section:
        flush_buffer(current_section, buffer)
        
    print(f"DEBUG: Parsed Sections: Summary={bool(data['summary'])}, Roles={len(data['responsibilities'])}, Edu={len(data['education'])}")

    return data

def build_resume_string(data):
    """
    Constructs the PLAIN TEXT resume string (No Markdown/JSON).
    This is what is sent to Flutter for the text preview.
    """
    # Safe Getters
    name = data.get("full_name") or data.get("name") or "Your Name"
    phone = data.get("phone") or ""
    email = data.get("email") or ""
    loc = data.get("location") or ""
    
    summary = data.get("summary") or "Professional Summary..."
    edu = data.get("education") or "Education Details..."
    roles = data.get("responsibilities") or "Roles..."
    proj = data.get("projects") or "Projects..."
    skills = data.get("skills") or "Skills..."
    certifications = data.get("certifications") or ""
    
    return f"""{name}
| {phone} | {email} | {loc} |

Professional Summary
{summary}

Education
{edu}

Roles & Responsibilities
{roles}

Projects
{proj}

Technologies & Tools
{skills}
"""

def build_resume_data(user: dict):
    """
    Standardizes user data into the fixed resume format.
    Handles missing keys gracefully.
    """
    # Helper to clean "None" strings and AI placeholders
    def clean_val(v, default=""):
        if not v: return default
        s = str(v).strip()
        if s.lower() in ["none", "null", "undefined"]: 
            return default
        # Only remove [Specify...] type placeholders, NOT user content in brackets
        if "[specify" in s.lower() or "[insert" in s.lower() or "[relevant" in s.lower() or "[year" in s.lower() or "[dates]" in s.lower():
            return default
        return s
        

    # 1. Extract/Derive fields for Summary
    # Try to find degree/specialization from education if not explicit
    degree = clean_val(user.get("degree"), "Degree")
    specialization = clean_val(user.get("specialization"), "Specialization")
    
    if (degree == "Degree" or specialization == "Specialization") and user.get("education"):
        # Try to infer from first education entry
        try:
            first_edu = user["education"][0]
            if isinstance(first_edu, dict):
                 # simplistic parse: "B.Tech in AI" -> deg="B.Tech", spec="AI"
                 raw_course = first_edu.get("course", "") or first_edu.get("degree", "")
                 if " in " in raw_course:
                     parts = raw_course.split(" in ")
                     degree = parts[0]
                     specialization = parts[1]
                 else:
                     degree = raw_course
        except:
            pass

    # Core skills (take first 3 of skills if not provided)
    core_skills = user.get("core_skills", [])
    if not core_skills and user.get("skills"):
        skills_val = user["skills"]
        if isinstance(skills_val, list):
            core_skills = skills_val[:3]
        elif isinstance(skills_val, str):
            # Split string by comma and take first 3
            core_skills = [s.strip() for s in skills_val.split(",")[:3]]
        else:
            core_skills = []
        
    # Ensure core_skills is always a list for joining
    if isinstance(core_skills, str):
        core_skills = [s.strip() for s in core_skills.split(",")]
        
    summary = clean_val(user.get("summary"))
    if not summary:
        skills_str = ', '.join(core_skills) if core_skills else "various technologies"
        role_label = clean_val(user.get("detected_role"), specialization or "Technical")
        summary = (
            f"Aspiring {role_label} professional and {degree} graduate in {specialization} with hands-on project experience in "
            f"{skills_str}. Proven ability to develop data-driven solutions, and a strong foundation in "
            "analytical problem-solving. Highly motivated to contribute to innovative projects and "
            "eager to grow in a dynamic professional environment."
        )

    # 2. Education section
    edu_list = user.get("education", [])
    formatted_edu = []
    if isinstance(edu_list, list):
        for edu in edu_list:
            if isinstance(edu, dict):
                inst = edu.get("institute") or edu.get("university", "")
                course = edu.get("course") or edu.get("degree", "")
                raw_grade = str(edu.get("cgpa") or edu.get("grade", ""))
                grade_str = f" | CGPA: {raw_grade}" if raw_grade else ""
                
                formatted_edu.append(f"{inst} – {course}{grade_str}")
            else:
                formatted_edu.append(str(edu))
        education = "\n".join(formatted_edu)
    else:
        education = str(edu_list)

    # 3. Responsibilities / Experience (Dynamic)
    # If the user has 'responsibilities' (from AI) or 'experience' (from parser), use it.
    # Otherwise use generic fresher text.
    
    raw_resp = user.get("responsibilities") or user.get("experience") or []
    
    if raw_resp:
        if isinstance(raw_resp, list):
            # If it's a list (from AI), format nicely key-points
            # Mistral might return dictionaries {role, company...} or strings
            formatted_resp = []
            for item in raw_resp:
                if isinstance(item, dict):
                    # Format: Role at Company (Date) \n - Details
                    role = item.get("role", "Role")
                    comp = item.get("company", "")
                    date = item.get("date", "")
                    details = item.get("details", [])
                    
                    header = f"• {role}"
                    if comp: header += f" at {comp}"
                    if date: header += f" ({date})"
                    
                    formatted_resp.append(header)
                    if isinstance(details, list):
                        for d in details:
                            formatted_resp.append(f"  - {d}")
                    else:
                        formatted_resp.append(f"  - {str(details)}")
                        
                else:
                    # Simple string point
                    formatted_resp.append(f"• {str(item)}")
            
            responsibilities = "\n".join(formatted_resp)
            
        elif isinstance(raw_resp, str):
            responsibilities = raw_resp
    
    # Only use professional fallback IF NO responsibilities were provided
    if not responsibilities:
        role_label = clean_val(user.get("detected_role"), specialization or "Software Professional")
        responsibilities = (
            f"• Technical Collaboration: Worked within project teams to identify and resolve technical challenges during development lifecycles.\n"
            f"• Solution Implementation: Applied {skills if skills else 'industry standard tools'} to develop, test, and optimize project-specific workflows and data pipelines.\n"
            f"• Quality Assurance: Conducted rigorous debugging and performance tuning to ensure high-quality delivery and project consistency.\n"
            f"• Documentation: Authored detailed technical specifications and project findings to ensure reproducibility and knowledge sharing."
        )

    # 4. Projects section
    proj_list = user.get("projects", [])
    formatted_proj = []
    if isinstance(proj_list, list):
        for proj in proj_list:
            if isinstance(proj, dict):
                title = proj.get("title") or proj.get("name", "")
                # Clean title markdown if present (e.g. **Title**)
                title = title.replace("**", "").replace("*", "").strip()
                
                desc = proj.get("description") or (proj.get("details", [])[0] if proj.get("details") else "")
                # Clean desc markdown if present
                desc = desc.replace("**", "").strip()

                formatted_proj.append(f"{title}\n• {desc}")
            else:
                formatted_proj.append(str(proj))
        projects = "\n\n".join(formatted_proj)
    else:
        projects = str(proj_list)

    # 5. Skills
    skills_list = user.get("skills", [])
    if isinstance(skills_list, list):
        skills = ", ".join(skills_list)
    else:
        skills = str(skills_list)

    # 6. Certifications
    cert_list = user.get("certifications", [])
    if isinstance(cert_list, list):
        formatted_cert = []
        for c in cert_list:
            formatted_cert.append(f"• {str(c)}")
        certifications = "\n".join(formatted_cert)
    else:
        certifications = str(cert_list)
        
    # Validation for Summary: If it's still generic/empty but we have scraped a role/skills, improve it
    if "graduate in" in summary and "Enthusiastic learner" in summary:
         pass # It's the default, maybe leave it or try to enrich if we have more info?

    return {
        "full_name": clean_val(user.get("name"), ""),
        "phone": clean_val(user.get("phone"), ""),
        "email": clean_val(user.get("email"), ""),
        "location": clean_val(user.get("location"), ""),
        "summary": summary,
        "education": education,
        "responsibilities": responsibilities,
        "projects": projects,
        "skills": skills,
        "certifications": certifications,
    }

class GenerateRequest(BaseModel):
    text: str
    name: str = ""
    email: str = ""
    phone: str = ""
    jobrole: str = ""
    template_id: str = "classic" 

class TargetedResumeRequest(BaseModel):
    personal_info: str
    job_description: str
    template_id: str = "classic"


# ================================
# [INPUT MODEL FOR PARSING]
# ================================
class VoiceText(BaseModel):
    text: str


# ================================
# [SAVE FUNCTION (SHEET + CSV FALLBACK)]
# ================================
def save_data(data):
    # 1. Try Google Sheets
    try:
        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        
        if os.path.exists("credentials.json"):
            creds = ServiceAccountCredentials.from_json_keyfile_name(
                "credentials.json", scope
            )
            client = gspread.authorize(creds)
            
            # ✅ Open by URL provided by user
            sheet_url = "https://docs.google.com/spreadsheets/d/1_AUHN66wwgHt48MVrc66uzzp8qdA1BTzTh0Gvn8aUpA/edit"
            sheet = client.open_by_url(sheet_url).sheet1
            
            sheet.append_row([
                data["name"],
                data["mailid"],
                data["phone"],
                data["jobrole"],
                str(datetime.now())
            ])
            return "[OK] Saved to Google Sheet"
        else:
            raise FileNotFoundError("credentials.json not found")
            
    except Exception as e:
        print(f"Google Sheet Error: {e}")
        
        # 2. Fallback to Local CSV
        csv_file = os.path.join(os.getcwd(), "leads_data.csv")
        file_exists = os.path.isfile(csv_file)
        
        timestamp = str(datetime.now())
        
        row_data = [
            data["name"],
            data["mailid"],
            data["phone"],
            data["jobrole"],
            timestamp
        ]

        # Retry mechanism: Try for 15 seconds
        max_retries = 15
        
        for attempt in range(max_retries):
            try:
                # Try appending to the main file
                with open(csv_file, "a", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    if not file_exists and attempt == 0: 
                        writer.writerow(["Name", "Email", "Phone", "Job Role", "Timestamp"])
                    writer.writerow(row_data)
                return "[OK] Saved to CSV"
            
            except PermissionError:
                # If locked, wait and retry
                if attempt < max_retries - 1:
                    time.sleep(1) # Wait 1 second
                    continue
                else:
                    return "[ERROR] FILE LOCKED. Please close 'leads_data.csv' immediately!"
            except Exception as e:
                return f"[ERROR] Save Failed: {str(e)}"


# ================================
# [TEXT PARSING FUNCTION]
# ================================
def extract_fields(text: str):
    # Try multiple patterns for name
    # 1. "Name [Name]" (like "name Cherry")
    # 2. "I am [Name]"
    # 3. "My name is [Name]"
    name_match = re.search(r"(?:my name is|this is|i am|name)\s+([a-zA-Z ]{1,30})", text, re.I)
    
    # Clean up name (stop at likely keywords if greedy match went too far)
    name = ""
    if name_match:
        raw_name = name_match.group(1).strip()
        # Split by known keywords to avoid capturing "Cherry education..."
        keywords = ["education", "job", "position", "applying", "role", "skill", "interest"]
        for kw in keywords:
            if f" {kw}" in raw_name.lower():
                raw_name = raw_name.lower().split(f" {kw}")[0]
        name = raw_name.title()

    email = re.search(r"[\w\.-]+@[\w\.-]+\.\w+", text)
    phone = re.search(r"\b\d{10}\b", text)

    job_patterns = [
        r"(data analyst|analyst)",
        r"(software engineer|developer)",
        r"(python developer)",
        r"(full stack developer)",
        r"(machine learning engineer)",
        r"(student)",
        r"(ui ux designer)",
        r"(tester)"
    ]

    jobrole = None
    for pattern in job_patterns:
        match = re.search(pattern, text, re.I)
        if match:
            jobrole = match.group(1)
            break

    return {
        "name": name if name else "",
        "mailid": email.group() if email else "",
        "phone": phone.group() if phone else "",
        "jobrole": jobrole.title() if jobrole else "",
        "raw_text": text
    }

def extract_resume_details(text: str):
    """Deep extraction for Resume specific fields"""
    base_info = extract_fields(text)
    
    # 1. Location
    loc_match = re.search(r"(?:from|in|at|located in)\s+([A-Z][a-z]+(?: [A-Z][a-z]+)*)", text)
    location = loc_match.group(1) if loc_match else "India"
    
    # 2. Education (Structured Extraction)
    education_list = []
    # Pattern: "B.Tech in AI ... at JNTU ... 2020-2024"
    # This is hard to regex perfectly, but let's try a few robust patterns
    
    # Try finding degree + college + year
    # Example: "B.Tech in AI from JNTU (2020 - 2024)"
    # Example: "B.Tech in AI from JNTU (2020 - 2024)"
    edu_matches = re.finditer(r"(B\.?Tech|Bachelor|Master|Degree|Diploma|SSC|HSC|Intermediate|Class\s*10|Class\s*12|Standard\s*10|Standard\s*12)\s+(?:in\s+)?([A-Za-z0-9 &]+?)\s+(?:from|at)\s+([A-Za-z0-9 \-\.,]+?)\s*((?:20\d\d)\s*-\s*(?:20\d\d|Present)?)", text, re.I)
    
    found_edu = False
    for m in edu_matches:
        found_edu = True
        education_list.append({
            "degree": f"{m.group(1)} {m.group(2)}",
            "university": m.group(3).strip(),
            "year": m.group(4).strip(),
            "grade": "" # Hard to extract reliably
        })
        
    if not found_edu:
         # Fallback to simple line detection if strictly structured regex fails
         # Look for lines containing "College" or "University"
         lines = text.split('\n')
         for line in lines:
             if "college" in line.lower() or "university" in line.lower() or "b.tech" in line.lower():
                 # Create a dummy structured object so it formats nicely
                 education_list.append({
                     "degree": line.strip(),
                     "university": "",
                     "year": "",
                     "grade": ""
                 })

    # 3. Projects
    # Look for "project on [Name]" or "worked on [Name]"
    projects_list = []
    # Split by "project" keyword to find segments
    project_segments = re.split(r"\bproject\b", text, flags=re.I)
    if len(project_segments) > 1:
        for seg in project_segments[1:]:
             # Take the first ~15 words as the project description/name
             words = seg.split()[:20]
             if words:
                # Name is first few words
                p_name = " ".join(words[:4]).title()
                p_desc = " ".join(words[4:]).capitalize() + "."
                projects_list.append({
                    "name": p_name,
                    "details": [p_desc]
                })
        
    # 4. Skills
    skills_keywords = ["python", "java", "sql", "excel", "html", "css", "javascript", "react", "flutter", "aws", "c++", "c", "machine learning", "ai", "communication", "teamwork"]
    found_skills = [sk.title() for sk in skills_keywords if sk in text.lower()]
    skills_text = ", ".join(found_skills) if found_skills else "Python, SQL, Excel"
    
    base_info.update({
        "location": location,
        "summary": text[:200].replace('\n', ' ') + "..." if len(text) > 50 else "Experienced professional looking for opportunities.",
        "education": education_list if education_list else [{"degree": "Degree Details", "university": "University Name", "year": "Year"}],
        "projects": projects_list if projects_list else "Project Experience available upon request.",
        "roles": [{"role": "Candidate", "company": "Previous Company", "date": "Dates", "details": ["Worked on key projects.", "Collaborated with team."]}],
        "skills": skills_text,
    })
    return base_info

# ================================
# ✅ ROBUST LOCAL FALLBACK
# ================================
def parse_cover_letter_locally(text: str):
    """
    Manually parses the structured resume text (Name: ..., Experience...) 
    and builds the Gold Standard Narrative Cover Letter.
    This GUARANTEES the format even if AI fails.
    """
    # 1. Extract Name
    # Try explicit "Name: ..." first, then first line
    name_match = re.search(r"(?:Name|Candidate Name):\s*(.*)", text, re.I)
    if name_match:
        name = name_match.group(1).strip()
    else:
        # Fallback: Assume the very first line is the name if it's short
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        name = lines[0] if lines and len(lines[0]) < 50 else "Your Name"

    # 2. Extract Role
    # Try explicit "Job Title...", then look for common keywords
    role_match = re.search(r"(?:Job Title|Role|Position):\s*(.*)", text, re.I)
    if role_match:
        role = role_match.group(1).strip()
    else:
        # Fallback: Look for keywords in the first few lines
        roles = ["Designer", "Developer", "Engineer", "Manager", "Analyst"]
        parsed_role = "Candidate"
        for line in text.split('\n')[:5]:
            for r in roles:
                if r.lower() in line.lower():
                    parsed_role = line.strip()
                    break
        role = parsed_role

    # 3. Extract Email/Phone (for footer)
    email_match = re.search(r"([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})", text)
    email = email_match.group(1).strip() if email_match else ""
    
    phone_match = re.search(r"(\+?\d[\d -]{8,15})", text)
    phone = phone_match.group(1).strip() if phone_match else ""
    
    # 4. Extract Companies (Smart Regex)
    companies = re.findall(r"(?:at|@)\s+([A-Za-z0-9 ]+?)\s+(?:\(|-|20)", text)
    if not companies:
         # Try generic company lookups
         companies = re.findall(r"([A-Z][a-zA-Z0-9 ]+ (?:Solutions|Technologies|Systems|Corp|Inc|LLC|Pvt|Ltd))", text)
    
    company_1 = companies[0].strip() if len(companies) > 0 else "my previous company"
    company_2 = companies[1].strip() if len(companies) > 1 else None
    
    # 5. Extract Skills/Tools
    tools_match = re.search(r"(?:Tools|Skills):\s*(.*)", text, re.I)
    tools = tools_match.group(1).strip() if tools_match else "standard industry tools"
    
    # --- BUILD THE NARRATIVE BODY ---
    
    # Para 1: Intro
    p1 = (f"I am writing to apply for the {role} position at your organization. "
          f"With my robust background in this field, I am confident in my ability to contribute to your team.")
    
    # Para 2: Background
    p2 = (f"My name is {name}, and I am a {role}. "
          f"My background includes designing user-centered interfaces and creating high-quality solutions.")
    
    # Para 3: Experience (The Story)
    if company_2:
        p3 = (f"In my recent role at {company_1}, I worked on core product tools and improved user experiences. "
              f"Prior to that, at {company_2}, I handled design and development tasks for various client projects.")
    else:
        p3 = (f"In my recent role at {company_1}, I worked on core product tools and improved user experiences, "
              f"collaborating closely with cross-functional teams.")
              
    # Para 4: Skills
    p4 = (f"I am proficient in tools such as {tools}. "
          f"I enjoy creating clean, user-friendly solutions and improving overall process efficiency.")
          
    # Para 5: Closing
    p5 = ("I am a quick learner, detail-oriented, and always willing to take on new challenges. "
          "I would be grateful for the opportunity to discuss how I can add value to your organization.")
          
    narrative_body = f"{p1}\n\n{p2}\n\n{p3}\n\n{p4}\n\n{p5}"
    
    return {
        "name": name,
        "email": email,
        "phone": phone,
        "jobrole": role,
        "body": narrative_body
    }

# ================================
# ✅ AI CONTENT GENERATOR (GEMINI & OPENAI)
# ================================


def generate_content_with_gemini(text: str, doc_type: str = "resume"):
    """
    Uses Google Gemini to extract and polish content.
    """
    try:
        model = genai.GenerativeModel('gemini-1.5-flash')
        
        if doc_type == "resume":
            prompt = (
                "You are a STRICT Resume Data Extractor.\n"
                "Task: Convert the raw user input into a clean, professional JSON structure for a resume.\n"
                "CRITICAL: Do NOT skip any details from the input. Extract EVERYTHING related to Name, Education, Projects, and Skills.\n"
                "STRICTLY follow this structure for the values:\n"
                "- summary: A professional summary (3-4 lines). Include degree and key skills.\n"
                "- education: List ALL degrees/schools found. Return list of objects: {university, degree, year, grade}.\n"
                "- roles: Professional Experience. Use bullet points.\n"
                "- projects: List ALL projects mentioned with 1-line description. Return list of objects: {name, details (list of strings)}.\n"
                "- skills: List ALL technical/soft skills mentioned.\n"
                "- name, email, phone, location: Extract accurately.\n"
                f"Raw Input: '{text}'"
            )
        else:
            prompt = (
                "You are an expert professional resume writer. Write a COVER LETTER that strictly follows this EXACT structure and tone:\n"
                "You MUST mention specific Company Names, Dates, and Projects found in the input.\n\n"
                "REFERENCE STRUCTURE (Follow this exactly):\n"
                "Paragraph 1: 'I am writing to express my interest in the [Role] position at your organization. With over [Years] of experience in [Field], along with strong hands-on expertise in [Key Skills], I am excited about the opportunity to contribute to your team.'\n\n"
                "Paragraph 2: 'My name is [Name], and I am a [Role] based in [Location]. My background includes [Brief Summary of Core Competencies].'\n\n"
                "Paragraph 3 (CRITICAL): 'In my recent role at [Company 1] ([Dates]), I [What you did]. Prior to that, at [Company 2] ([Dates]), I [What you did]. I also bring experience from [Company 3] where I [What you did].' (If only 1 company, expand on it).\n\n"
                "Paragraph 4: 'I am proficient in tools such as [Tools List] and experienced in [Skills List]. Some of my project experience includes [Project List].'\n\n"
                "Paragraph 5: 'Along with technical skills, I am [Soft Skills]. I am eager to bring my creativity and problem-solving ability to your organization.'\n\n"
                "Paragraph 6: 'Thank you for considering my application...'\n\n"
                "Do NOT include the header (Dear Hiring Manager) or footer (Sincerely) in the 'body' field. Just the paragraphs.\n"
                "Return valid JSON with keys: name, email, phone, jobrole, body.\n"
                f"Raw Input: '{text}'"
            )
            
        # Helper to ensure JSON response
        response = model.generate_content(prompt, generation_config={"response_mime_type": "application/json"})
        
        # Clean potential markdown wrapping
        content = response.text.strip()
        if content.startswith("```json"):
            content = content[7:-3]
        
        print(f"Gemini Response: {repr(content)[:200]}...")
        return json.loads(content)
    except Exception as e:
        print(f"[ERROR] Gemini Error: {e}")
        return None

def generate_content_ai(text: str, doc_type: str = "resume"):
    """
    Wrapper to choose between Gemini and OpenAI.
    """
    # Check for Python
    if os.path.exists("D:\\Python314\\python.exe"):
        python_exe = "D:\\Python314\\python.exe"
    else:
        python_exe = "python"

    # User Request: Force OpenAI for Resume if key is available
    if doc_type == "resume" and openai_api_key:
        print("Using OpenAI for Resume generation...")
        return generate_content_with_openai(text, doc_type)
        
    # Prefer Gemini for Cover Letters or if OpenAI is missing
    if gemini_api_key:
        print("Using Gemini...")
        return generate_content_with_gemini(text, doc_type)
    elif openai_api_key:
        print("Using OpenAI...")
        return generate_content_with_openai(text, doc_type)
    else:
        return None

def generate_content_with_openai(text: str, doc_type: str = "resume"):
    """
    Uses OpenAI to extract and polish content from raw speech.
    doc_type: 'resume' or 'cover_letter'
    """
    if doc_type == "resume":
        prompt = (
            "You are a SENIOR RESUME ARCHITECT. Your goal is to turn sparse user input into a PREMIER, ATS-OPTIMIZED resume.\n\n"
            "STRICT RULES:\n"
            "1. NO PLACEHOLDERS: Do NOT use '[Placeholders]', '[Dates]', or '[Institution Name]'. If a value is missing, infer a professional, realistic description or use general elite terminology.\n"
            "2. NARRATIVE EXPANSION: Intelligently expand fragments into professional bullet points. If they say 'I know Python', expand to 'Software Engineering: Leveraged Python for automated data processing and algorithm implementation.'\n"
            "3. STYLE: Match the tone of a high-achieving B.Tech graduate (like Deepti Challapalli). Use powerful action verbs.\n"
            "4. ROLES: If the user is a fresher, populate 'roles_and_responsibilities' with high-impact descriptions of their technical training and project teamwork.\n"
            "5. NO HALLUCINATION OF COMPANIES: Do not invent fake company names, but DO invent professional descriptions of their skills and achievements.\n\n"
            "Return ONLY valid JSON in the following structure:\n"
            "{\n"
            '  "name": "Full Name",\n'
            '  "email": "Professional Email",\n'
            '  "phone": "Phone Number",\n'
            '  "location": "City, State, Country",\n'
            '  "detected_role": "Target Job Title",\n'
            '  "professional_summary": "3-4 lines of high-impact narrative summary...",\n'
            '  "education": [ "Detailed Education Entry (No placeholders)" ],\n'
            '  "roles_and_responsibilities": [ "Significant achievement 1", "Technical responsibility 2" ],\n'
            '  "projects": [ "**Project Title:** Detailed description of technical implementation and results." ],\n'
            '  "technical_skills": [ "Skill Name" ]\n'
            "}\n\n"
            f"User Information:\n<<< {text} >>>"
        )
    else:
        prompt = (
            "You are an expert professional resume writer. Write a COVER LETTER that strictly follows this EXACT structure and tone:\n"
            "You MUST mention specific Company Names, Dates, and Projects found in the input.\n\n"
            "REFERENCE STRUCTURE (Follow this exactly):\n"
            "Paragraph 1: 'I am writing to express my interest in the [Role] position at your organization. With over [Years] of experience in [Field], along with strong hands-on expertise in [Key Skills], I am excited about the opportunity to contribute to your team.'\n\n"
            "Paragraph 2: 'My name is [Name], and I am a [Role] based in [Location]. My background includes [Brief Summary of Core Competencies].'\n\n"
            "Paragraph 3 (CRITICAL): 'In my recent role at [Company 1] ([Dates]), I [What you did]. Prior to that, at [Company 2] ([Dates]), I [What you did]. I also bring experience from [Company 3] where I [What you did].' (If only 1 company, expand on it).\n\n"
            "Paragraph 4: 'I am proficient in tools such as [Tools List] and experienced in [Skills List]. Some of my project experience includes [Project List].'\n\n"
            "Paragraph 5: 'Along with technical skills, I am [Soft Skills]. I am eager to bring my creativity and problem-solving ability to your organization.'\n\n"
            "Paragraph 6: 'Thank you for considering my application...'\n\n"
            "Do NOT include the header (Dear Hiring Manager) or footer (Sincerely) in the 'body' field. Just the paragraphs.\n"
            "Return JSON with keys: name, email, phone, jobrole, body.\n"
            f"Raw Input: '{text}'"
        )

    # 1. Try Mistral AI (First Priority as requested)
    mistral_api_key = os.environ.get("MISTRAL_API_KEY", "ZVzUQWNuusFD28EPJZia3E6ecLrgC1Em") # User provided key
    if mistral_api_key:
        try:
            from mistralai import Mistral
            client = Mistral(api_key=mistral_api_key)
            
            # Use Mistral Large or Small based on complexity
            model = "mistral-large-latest" 
            
            chat_response = client.chat.complete(
                model=model,
                messages=[
                    {
                        "role": "system",
                        "content": "You are a helpful assistant that outputs only valid structured JSON."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    },
                ],
                response_format={"type": "json_object"}
            )
            
            content = chat_response.choices[0].message.content
            print(f"Mistral Response: {repr(content)[:100]}...")
            return json.loads(content)
        except Exception as e:
            print(f"Mistral Error: {e}")
            # Fall through to OpenAI
            
    # 2. Try OpenAI (Second Priority)
    if openai.api_key:
        try:
            client = openai.OpenAI()
            response = client.chat.completions.create(
                model="gpt-4o-mini", 
                messages=[
                    {"role": "system", "content": "You are a helpful assistant that outputs only valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.7,
                response_format={"type": "json_object"}
            )
            
            content = response.choices[0].message.content
            print(f"OpenAI Response: {repr(content)[:100]}...")
            return json.loads(content)
        except Exception as e:
            print(f"OpenAI Error: {str(e)}")
            
    # 3. Try Gemini (Fallback)
    # ... (existing gemini code would go here if not removed, or return None)
    return None


# ================================
# TRANSLATION HELPER
# ================================


def translate_to_english(text: str) -> str:
    """Translate text to English if it's in another language"""
    try:
        # Detect language
        # Detect is hard with standard deep_translator for free, 
        # so we just force translate to 'en' from 'auto'
        translator = GoogleTranslator(source='auto', target='en')
        translated = translator.translate(text)
        return translated
    except Exception as e:
        print(f"Translation Error: {e}")
        return text


# ================================
# API: PARSE TEXT AND SAVE TO EXCEL
# ================================
@app.post("/parse-and-save")
async def parse_and_save(data: VoiceText):
    try:
        # Log the received text for debugging
        print(f"\nReceived text from Flutter: {data.text}")
        
        # Translate to English if needed
        english_text = translate_to_english(data.text)
        print(f"Text for parsing: {english_text}")
        
        # Extract fields from the translated English text
        parsed = extract_fields(english_text)
        print(f"Parsed fields: {parsed}")
        
        # Save to Excel file
        excel_file = os.path.join(os.getcwd(), "leads_data.xlsx")
        
        timestamp = str(datetime.now())
        
        # Row data: Name, Email, Phone, Job Role, Timestamp
        row_data = [
            parsed["name"],
            parsed["mailid"],
            parsed["phone"],
            parsed["jobrole"],
            timestamp
        ]

        # Retry mechanism for file locking
        max_retries = 15
        
        for attempt in range(max_retries):
            try:
                # Check if file exists
                if os.path.isfile(excel_file):
                    # Load existing workbook
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_file)
                    ws = wb.active
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    # Write header
                    ws.append(["Name", "Email", "Phone", "Job Role", "Timestamp"])
                
                # Append data row
                ws.append(row_data)
                
                # Save workbook
                wb.save(excel_file)
                
                print(f"Saved to Excel: {row_data}")
                
                return {
                    "status": f"Saved to {excel_file}",
                    "data": parsed
                }
            
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(1)
                    continue
                else:
                    return {
                        "status": "[ERROR] FILE LOCKED. Close leads_data.xlsx!",
                        "error": "File is locked"
                    }
            except Exception as e:
                return {
                    "status": f"[ERROR] Save Failed: {str(e)}",
                    "error": str(e)
                }


    except Exception as e:
        print(f"[ERROR] Error: {str(e)}")
        return {
            "status": "[ERROR] Failed",
            "error": str(e)
        }




# ================================
# WEBSOCKET ENDPOINT
# ================================
@app.websocket("/ws/transcribe")
async def ws_transcribe(websocket: WebSocket):
    """
    Client sends raw audio bytes (webm) in chunks.
    Server runs Whisper on the accumulated audio and sends
    interim text back after each chunk.
    """
    await websocket.accept()
    audio_buffer = bytearray()
    last_chunk_time = asyncio.get_event_loop().time()

    try:
        async for message in websocket:
            # `message` is a binary chunk from the client
            if isinstance(message, bytes):
                audio_buffer.extend(message)
                # If we haven't received a chunk for >1.5s, treat it as end-of-speech
                now = asyncio.get_event_loop().time()
                if now - last_chunk_time > 1.5:
                    # Run Whisper on the whole buffer
                    if len(audio_buffer) > 0:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".webm") as tmp:
                            tmp.write(audio_buffer)
                            tmp_path = tmp.name

                        try:
                            # Use the same prompt as before
                            prompt = "The following is a clear conversation. Please transcribe it accurately."
                            result = model.transcribe(tmp_path, initial_prompt=prompt)
                            if result["text"].strip():
                                await websocket.send_text(result["text"])
                        except Exception as e:
                            print(f"Transcription error: {e}")
                        finally:
                            if os.path.exists(tmp_path):
                                os.unlink(tmp_path)
                        
                        # Reset for next utterance
                        audio_buffer.clear()
                last_chunk_time = now
            else:
                # ignore non-binary messages
                continue
    except ConnectionClosedOK:
        print("WebSocket closed normally")
    except Exception as e:
        print(f"WebSocket error: {e}")
import io
from fastapi import Body
from fastapi.responses import StreamingResponse
# ================================
# NEW IMPORTS FOR PDF GENERATION
# ================================
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import simpleSplit

def _create_text_file(content: str, filename: str) -> StreamingResponse:
    """Create a StreamingResponse for a plain text file download."""
    file_like = io.BytesIO(content.encode("utf-8"))
    return StreamingResponse(
        file_like,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )

# ================================
# PDF CREATION HELPER
# ================================
# ================================
# PROFESSIONAL TEMPLATE ENGINE (PLATYPUS)
# ================================
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib import colors

def _create_platypus_pdf(elements, filename: str) -> StreamingResponse:
    """Builds a PDF from a list of Flowables (Paragraphs, Tables, etc.)"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=50, leftMargin=50, 
        topMargin=50, bottomMargin=50
    )
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="{filename}"'
    })

def _get_styles():
    styles = getSampleStyleSheet()
    # reliable styles
    styles.add(ParagraphStyle(name='Justify', parent=styles['Normal'], alignment=TA_JUSTIFY, spaceAfter=6))
    styles.add(ParagraphStyle(name='Center', parent=styles['Normal'], alignment=TA_CENTER, spaceAfter=6))
    styles.add(ParagraphStyle(name='HeaderName', parent=styles['Normal'], alignment=TA_CENTER, fontSize=24, fontName='Helvetica-Bold', spaceAfter=12))
    styles.add(ParagraphStyle(name='HeaderContact', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10, fontName='Helvetica', spaceAfter=20, textColor=colors.darkgrey))
    styles.add(ParagraphStyle(name='SectionHeader', parent=styles['Normal'], fontSize=14, fontName='Helvetica-Bold', spaceAfter=6, spaceBefore=12, borderPadding=2, borderColor=colors.black, borderWidth=0, borderRadius=0))
    # We will simulate a line under the section header using a table or specific paragraph style, 
    # but strictly speaking, a simple Bold Header is the standard "Template" look.
    return styles

@app.post("/generate-cover-letter-pdf")
async def generate_cover_letter_pdf(req: GenerateRequest):
    """Generate a cover letter using a standard Professional Template (AI Enhanced)."""
    
    # Try AI (Gemini or OpenAI) first
    ai_data = generate_content_ai(req.text, "cover_letter")
    
    if ai_data:
        # AI success
        name = ai_data.get("name", req.name or "Your Name")
        email = ai_data.get("email", req.email or "")
        phone = ai_data.get("phone", req.phone or "")
        jobrole = ai_data.get("jobrole", req.jobrole or "Candidate")
        body_content = ai_data.get("body", req.text)
    else:
        # FALLBACK: Use Robust Parsing (No more raw dumps)
        local_data = parse_cover_letter_locally(req.text)
        name = local_data["name"]
        email = local_data["email"]
        phone = local_data["phone"]
        jobrole = local_data["jobrole"]
        body_content = local_data["body"]
    
    styles = _get_styles()
    story = []

    # --- HEADER ---
    story.append(Paragraph(name, styles['HeaderName']))
    contact_info = f"{email} | {phone}"
    story.append(Paragraph(contact_info, styles['HeaderContact']))
    
    story.append(Spacer(1, 25))
    
    # --- DATE & RECIPIENT ---
    date_str = datetime.now().strftime("%B %d, %Y")
    story.append(Paragraph(date_str, styles['Normal']))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Hiring Manager", styles['Normal']))
    story.append(Paragraph("Recruitment Team", styles['Normal']))
    story.append(Spacer(1, 15))
    
    # --- SALUTATION ---
    story.append(Paragraph("Dear Hiring Manager,", styles['Normal']))
    story.append(Spacer(1, 10))
    
    # --- BODY ---
    for para in body_content.split('\n\n'):
        if para.strip():
            story.append(Paragraph(para.strip(), styles['Justify']))
            story.append(Spacer(1, 8))
            
    story.append(Spacer(1, 15))
    
    # --- SIGN OFF ---
    story.append(Paragraph("Sincerely,", styles['Normal']))
    story.append(Spacer(1, 20)) # Space for signature
    story.append(Paragraph(name, styles['Normal']))
    if email:
        story.append(Paragraph(f"Email: {email}", styles['Normal']))
    if phone:
        story.append(Paragraph(f"Phone: {phone}", styles['Normal']))
    
    return _create_platypus_pdf(story, 'cover_letter.pdf')

@app.post("/generate-resume-pdf")
async def generate_resume_pdf(req: GenerateRequest):
    """
    Download gateway: Ensures 1:1 match with UI content.
    """
    try:
        # Normalize text for checking
        low_text = req.text.lower()
        
        # If it looks like our structured UI output, use LITERAL PARSING
        if "|" in req.text and ("professional summary" in low_text or "summary" in low_text):
            print("[INFO] Download Clicked: Using Literal Sync.")
            parsed = parse_formatted_resume(req.text)
            
            # Map EXACTLY to what Template 45 expects
            final_data = {
                "name": parsed.get("name") or req.name,
                "full_name": parsed.get("name") or req.name,
                "phone": parsed.get("phone") or req.phone,
                "email": parsed.get("email") or req.email,
                "location": parsed.get("location") or "India",
                "summary": parsed.get("summary"),
                "education": parsed.get("education"),
                "responsibilities": parsed.get("responsibilities"),
                "projects": parsed.get("projects"),
                "skills": parsed.get("skills"),
                "certifications": parsed.get("certifications")
            }
            
            if req.template_id == "template_45":
                return _create_modern_resume(final_data, "resume.pdf")
            else:
                return _create_classic_pdf_internal(final_data)

        # AI Generation Path (for voice commands)
        ai_data = generate_content_ai(req.text, "resume")
        final_data = build_resume_data(ai_data if ai_data else {})
        # Merge request metadata if AI missed it
        if not final_data.get("full_name"): final_data["full_name"] = req.name
        
        if req.template_id == "template_45":
            return _create_modern_resume(final_data, "resume.pdf")
        else:
            return _create_classic_pdf_internal(final_data)

    except Exception as e:
        import traceback
        print(f"[ERROR] PDF ERROR: {traceback.format_exc()}")
        return {"error": str(e)}

    except Exception as e:
        print(f"[ERROR] Resume PDF Error: {traceback.format_exc()}")
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=500, content={"error": str(e), "detail": traceback.format_exc()})

def _create_classic_pdf_internal(final_data):
    """
    1:1 Sync Template: Matches the user's minimalist original (Image 1).
    Centered Header, Left-Aligned Body, No Lines, Helvetica Font.
    """
    styles = _get_styles()
    
    # Custom Styles for Exact Match
    name_style = ParagraphStyle('NameStyle', parent=styles['Normal'], fontSize=16, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=2)
    contact_style = ParagraphStyle('ContactStyle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, fontName='Helvetica', spaceAfter=20)
    section_style = ParagraphStyle('SecStyle', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', spaceBefore=15, spaceAfter=8, textTransform='uppercase')
    body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontSize=10, leading=13, alignment=TA_LEFT)
    bullet_style = ParagraphStyle('BulletStyle', parent=body_style, leftIndent=15, firstLineIndent=0, spaceAfter=4)
    
    story = []
    
    # 1. HEADER (Centered)
    story.append(Paragraph(final_data["full_name"], name_style))
    contact = " | ".join(filter(None, [final_data.get('phone'), final_data.get('email'), final_data.get('location')]))
    if contact:
        story.append(Paragraph(f"| {contact} |", contact_style))
    
    # 2. SECTIONS (Left Aligned, No Lines)
    def add_section(title, content):
        if not content: return
        story.append(Paragraph(title, section_style))
        
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if not line: continue
            
            # Format: Detect bullets or bold fragments
            formatted = line.replace("**", "<b>").replace("**", "</b>") # Basic bold swap
            
            if line.startswith('•') or line.startswith('-') or line.startswith('*'):
                # Clean prefix and use bullet style
                clean_line = re.sub(r'^[\s•\-\*]+', '', line)
                story.append(Paragraph(f"• {clean_line}", bullet_style))
            else:
                story.append(Paragraph(formatted, body_style))
        story.append(Spacer(1, 5))

    add_section("PROFESSIONAL SUMMARY", final_data.get("summary"))
    add_section("ACADEMIC PROFILE", final_data.get("education"))
    add_section("WORK EXPERIENCE", final_data.get("responsibilities"))
    add_section("PROJECTS", final_data.get("projects"))
    add_section("TECHNOLOGIES & TOOLS", final_data.get("skills"))
    if final_data.get("certifications"):
        add_section("CERTIFICATIONS", final_data.get("certifications"))
    
    return _create_platypus_pdf(story, 'resume.pdf')


def _create_modern_resume(data, filename):
    """
    Template 45 Style: Modern Blue Header with Clean Layout
    Optimized for correct formatting from build_resume_data.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    story = []
    
    # Custom Styles
    # Blue header bar style
    header_color = colors.HexColor("#2C3E50") # Dark Blue
    text_white = colors.white
    
    # --- HEADER SECTION (Name + Role) ---
    name = data.get("full_name", "Your Name").upper()
    
    # Table for Header
    header_style = ParagraphStyle('ModernHeader', parent=styles['Normal'], fontSize=24, textColor=text_white, alignment=TA_CENTER, fontName='Helvetica-Bold')
    contact_style = ParagraphStyle('ModernContact', parent=styles['Normal'], fontSize=10, textColor=text_white, alignment=TA_CENTER, fontName='Helvetica')
    
    contact_text = f"{data.get('phone','')} | {data.get('email','')} | {data.get('location','')}"
    
    p_name = Paragraph(name, header_style)
    p_contact = Paragraph(contact_text, contact_style)
    
    # 2-Row Table with Blue Background
    header_data = [[p_name], [p_contact]]
    t_header = Table(header_data, colWidths=['100%'])
    t_header.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), header_color),
        ('TOPPADDING', (0,0), (-1,-1), 20),
        ('BOTTOMPADDING', (0,0), (-1,-1), 20),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    
    story.append(t_header)
    story.append(Spacer(1, 20))
    
    # --- CONTENT STYLES ---
    section_head_style = ParagraphStyle('ModernSection', parent=styles['Normal'], fontSize=14, textColor=header_color, fontName='Helvetica-Bold', spaceBefore=15, spaceAfter=5)
    body_style = ParagraphStyle('ModernBody', parent=styles['Normal'], fontSize=11, leading=14)
    # Style for Bold Text in Body (if any)
    
    def clean_text_for_pdf(text):
        """Standardizes text, cleaning markdown to HTML for ReportLab if needed."""
        if not text: return ""
        # Remove markdown bold/italics markers if they exist (ReportLab needs <b> or nothing)
        # We did some cleaning in build_resume_data, but let's be safe.
        # Replace **text** with <b>text</b>
        t = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', text)
        t = re.sub(r'\*(.*?)\*', r'<i>\1</i>', t)
        return t

    # Helper to clean and split content
    def add_modern_section(title, content):
        if not content: return
        story.append(Paragraph(title.upper(), section_head_style))
        # Add a thin line under header
        story.append(Table([['']], colWidths=['100%'], rowHeights=[1], style=TableStyle([('LINEBELOW', (0,0), (-1,-1), 1, colors.lightgrey)])))
        story.append(Spacer(1, 8))
        
        # Parse content lines
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if not line: continue
            
            # Apply formatting
            formatted_line = clean_text_for_pdf(line)
            
            # Check for bullets
            if line.startswith('•') or line.startswith('-'):
                # Indented bullet
                bullet_style = ParagraphStyle('Bullet', parent=body_style, leftIndent=15, firstLineIndent=0)
                story.append(Paragraph(formatted_line, bullet_style))
            # Try to detect Titles (formatted like "Title" and next line is bullet) - Heuristic
            # But relying on Bold tag logic from clean_text is better
            else:
                # If it's a project title, we might have bolded it manually or cleaned it
                # Make lines without bullets slightly distinct if they look like headers?
                # No, just standard body style unless it has <b> tags.
                story.append(Paragraph(formatted_line, body_style))
        
        story.append(Spacer(1, 10))

    # Add Sections
    add_modern_section("Professional Summary", data.get("summary", ""))
    add_modern_section("Education", data.get("education", ""))
    add_modern_section("Roles & Responsibilities", data.get("responsibilities", ""))
    add_modern_section("Projects", data.get("projects", ""))
    add_modern_section("Technologies & Tools", data.get("skills", ""))
    add_modern_section("Certifications", data.get("certifications", ""))
    
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="{filename}"'
    })


    # --- HELPER: ADD SIMPLE TEXT BLOCKS (Handles newlines from build_resume_data) ---
    def add_block(text_content):
        # Split by newlines so each line is a Paragraph (better for bullets usually)
        # build_resume_data returns joined strings.
        # formatting there includes bullets `•`.
        for line in text_content.split('\n'):
            if line.strip():
                # Check for bullet
                if line.strip().startswith("•"):
                    # Use a bullet style or just text
                    story.append(Paragraph(line.strip(), styles['Normal']))
                else:
                    # Bold usage in text? The build_resume_data returns plain text "Title\n• Desc".
                    # We can try to bold the line if it doesn't start with bullet and is short?
                    # For now just print it.
                    story.append(Paragraph(line.strip(), styles['Normal']))
        story.append(Spacer(1, 8))
    
    # 1. PROFESSIONAL SUMMARY
    if final_data["summary"]:
        add_section_header("PROFESSIONAL SUMMARY")
        story.append(Paragraph(final_data["summary"], styles['Normal']))
        story.append(Spacer(1, 10))
        
    # 2. EDUCATION
    if final_data["education"]:
        add_section_header("EDUCATION")
        add_block(final_data["education"])

    # 3. ROLES & RESPONSIBILITIES
    if final_data["responsibilities"]:
        add_section_header("ROLES & RESPONSIBILITIES")
        add_block(final_data["responsibilities"])

    # 4. PROJECTS
    if final_data["projects"]:
        add_section_header("PROJECTS")
        add_block(final_data["projects"])

    # 5. TECHNOLOGIES & TOOLS
    if final_data["skills"]:
        add_section_header("TECHNOLOGIES & TOOLS")
        story.append(Paragraph(final_data["skills"], styles['Normal']))
        story.append(Spacer(1, 8))
        
    return _create_platypus_pdf(story, 'resume.pdf')


@app.post("/generate-cover-letter-pdf")
async def generate_cover_letter_pdf(req: GenerateRequest):
    """Generate a cover letter using a standard Professional Template (AI Enhanced)."""
    
    # Try AI (Gemini or OpenAI) first
    ai_data = generate_content_ai(req.text, "cover_letter")
    
    if ai_data:
        # AI success
        name = ai_data.get("name", req.name or "Your Name")
        email = ai_data.get("email", req.email or "")
        phone = ai_data.get("phone", req.phone or "")
        jobrole = ai_data.get("jobrole", req.jobrole or "Candidate")
        body_content = ai_data.get("body", req.text)
    else:
        # Fallback to local parsing
        if not req.name or not req.email:
             english_text = translate_to_english(req.text)
             parsed = extract_fields(english_text)
             if not req.name: req.name = parsed.get("name", "Your Name")
             if not req.email: req.email = parsed.get("mailid", "")
             if not req.phone: req.phone = parsed.get("phone", "")
             if not req.jobrole: req.jobrole = parsed.get("jobrole", "Candidate")
        
        name = req.name
        email = req.email
        phone = req.phone
        jobrole = req.jobrole
        # Default simple body
        body_content = (
            f"I am writing to express my enthusiastic interest in the {jobrole} position. "
            f"\n\n{req.text}\n\n"
            "Thank you for your time and consideration."
        )
    
    styles = _get_styles()
    story = []

    # --- HEADER ---
    story.append(Paragraph(name, styles['HeaderName']))
    contact_info = f"{email} | {phone}"
    story.append(Paragraph(contact_info, styles['HeaderContact']))
    
    story.append(Spacer(1, 25))
    
    # --- DATE & RECIPIENT ---
    date_str = datetime.now().strftime("%B %d, %Y")
    story.append(Paragraph(date_str, styles['Normal']))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Hiring Manager", styles['Normal']))
    story.append(Paragraph("Recruitment Team", styles['Normal']))
    story.append(Spacer(1, 15))
    
    # --- SALUTATION ---
    story.append(Paragraph("Dear Hiring Manager,", styles['Normal']))
    story.append(Spacer(1, 10))
    
    # --- BODY ---
    for para in body_content.split('\n\n'):
        if para.strip():
            story.append(Paragraph(para.strip(), styles['Justify']))
            story.append(Spacer(1, 8))
            
    story.append(Spacer(1, 15))
    
    # --- SIGN OFF ---
    story.append(Paragraph("Sincerely,", styles['Normal']))
    story.append(Spacer(1, 20)) # Space for signature
    story.append(Paragraph(name, styles['Normal']))
    if email:
        story.append(Paragraph(f"Email: {email}", styles['Normal']))
    if phone:
        story.append(Paragraph(f"Phone: {phone}", styles['Normal']))
    
    return _create_platypus_pdf(story, 'cover_letter.pdf')




@app.post("/download-cover-letter")
async def download_cover_letter(data: dict = Body(...)):
    raw_text = data.get('text', '')
    english_text = translate_to_english(raw_text)
    cover_letter = (
        "Dear Hiring Manager,\n\n"
        f"I am excited to apply for the position. {english_text}\n\n"
        "Sincerely,\nYour Name"
    )
    return _create_text_file(cover_letter, 'cover_letter.txt')

@app.post("/download-resume")
async def download_resume(data: dict = Body(...)):
    raw_text = data.get('text', '')
    english_text = translate_to_english(raw_text)
    resume = (
        "--- Resume ---\n\n"
        f"Personal Statement:\n{english_text}\n\n"
        "Experience:\n- ... (add your experience here)\n\n"
        "Education:\n- ... (add your education here)\n\n"
        "Skills:\n- ... (add your skills here)\n\n"
        "References available upon request."
    )
    return _create_text_file(resume, 'resume.txt')

@app.post("/generate-cover-letter-json")
async def generate_cover_letter_json(data: dict = Body(...)):
    raw_text = data.get('text', '')
    
    # Use AI for generation
    ai_data = generate_content_ai(raw_text, "cover_letter")
    
    if ai_data:
        name = ai_data.get("name", "Your Name")
        body = ai_data.get("body", "I am writing to express my interest...")
        
        cover_letter = (
            "Dear Hiring Manager,\n\n"
            f"{body}\n\n"
            f"Sincerely,\n{name}"
        )
    else:
        # [FALLBACK]: Use Robust Parsing
        local_data = parse_cover_letter_locally(raw_text)
        name = local_data["name"]
        
        cover_letter = (
            "Dear Hiring Manager,\n\n"
            f"{local_data['body']}\n\n"
            f"Sincerely,\n{name}\n{local_data['email']}\n{local_data['phone']}"
        )
        
    return {"cover_letter": cover_letter}

@app.post("/chat-resume")
async def chat_resume(data: dict = Body(...)):
    """
    Conversational endpoint for the 'Write to Resume' chat mode.
    Receives user message + chat history, returns only an AI follow-up question/message.
    Does NOT generate the resume — that is triggered separately.
    """
    user_message = data.get('message', '')
    history = data.get('history', [])  # List of {role, text}

    # Build conversation string for the AI
    history_str = "\n".join([f"{m['role'].upper()}: {m['text']}" for m in history])

    system_prompt = (
        "You are a friendly, professional AI Resume Assistant. "
        "The user is providing information for their resume. "
        "Your ONLY job is to acknowledge their input enthusiastically and tell them you are updating their live resume preview. "
        "Do NOT ask any follow-up questions. "
        "Keep your response to 1-2 short, encouraging sentences. "
        "Example: 'Got it! I've updated your resume with these details on the right.' or 'Perfect, adding that to your professional summary now!'"
    )

    prompt = f"{system_prompt}\n\nConversation so far:\n{history_str}\nUSER: {user_message}\nAI:"

    # Try Gemini first (free)
    try:
        gemini_model = genai.GenerativeModel('gemini-1.5-flash')
        response = gemini_model.generate_content(prompt)
        ai_reply = response.text.strip()
        ready = "[READY_TO_GENERATE]" in ai_reply
        ai_reply = ai_reply.replace("[READY_TO_GENERATE]", "").strip()
        return {"message": ai_reply, "ready": ready}
    except Exception as e:
        print(f"[chat-resume] Gemini error: {e}")

    # Fallback: Mistral
    mistral_api_key = os.environ.get("MISTRAL_API_KEY", "ZVzUQWNuusFD28EPJZia3E6ecLrgC1Em")
    if mistral_api_key:
        try:
            from mistralai import Mistral
            client = Mistral(api_key=mistral_api_key)
            chat_response = client.chat.complete(
                model="mistral-small-latest",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"Conversation so far:\n{history_str}\nUSER: {user_message}"}
                ]
            )
            ai_reply = chat_response.choices[0].message.content.strip()
            ready = "[READY_TO_GENERATE]" in ai_reply
            ai_reply = ai_reply.replace("[READY_TO_GENERATE]", "").strip()
            return {"message": ai_reply, "ready": ready}
        except Exception as e:
            print(f"[chat-resume] Mistral error: {e}")

    # Final fallback
    return {
        "message": "Got it! Could you tell me more about your skills and technologies you work with?",
        "ready": False
    }


@app.post("/generate-resume-json")
async def generate_resume_json(data: dict = Body(...)):
    raw_text = data.get('text', '')
    english_text = translate_to_english(raw_text)
    
    # [NEW] STRATEGY: PREFER LOCAL PARSING FOR STRUCTURED TEXT (Copy-Paste)
    # If the user pasted a resume with headers, they want THAT exact content formatted.
    # We check if it looks like our "Formatted Resume" (Name \n | Contact | \n Summary...)
    
    lines = [l.strip() for l in english_text.split('\n') if l.strip()]
    # Check for Pipe Separators in first few lines (Contact line)
    has_pipes = False
    for i in range(min(5, len(lines))):
        if "|" in lines[i] and len(lines[i]) > 10:
            has_pipes = True
            break
            
    if (has_pipes or "professional summary" in english_text.lower()) and len(lines) > 15:
        print("[INFO] Detected Full Formatted Resume Input. Using Direct Parsing.")
        parsed_data = parse_formatted_resume(english_text)
        
        # Mapping parse_formatted_resume output (which returns final keys) to final_data
        # parse_formatted_resume returns: name, phone, email, location, summary, education, responsibilities, projects, skills
        # All as STRINGS (joined sections).
        
        # We need to ensure we map keys correctly for _create_modern_resume
        # _create_modern_resume expects: full_name, phone, email, location, summary, education, responsibilities, projects, skills
        
        if parsed_data.get("name") or parsed_data.get("education"):
             final_data = {
                "full_name": parsed_data.get("name", "Your Name"),
                "phone": parsed_data.get("phone", ""),
                "email": parsed_data.get("email", ""),
                "location": parsed_data.get("location", ""),
                "summary": parsed_data.get("summary", ""),
                "education": parsed_data.get("education", ""),
                "responsibilities": parsed_data.get("responsibilities", ""),
                "projects": parsed_data.get("projects", ""),
                "skills": parsed_data.get("skills", ""),
                "certifications": parsed_data.get("certifications", ""),
            }
             # Post-process: Remove "None" string artifacts if any slipped through
             for k, v in final_data.items():
                 if str(v).strip().lower() == "none": final_data[k] = ""

             return {"resume": build_resume_string(final_data)}

    # If not formatted, check for semi-structured headers
    headers = ["education", "experience", "projects", "skills", "summary", "my projects", "work experience", "professional summary"]
    lower_text = english_text.lower()
    has_headers = any(h in lower_text for h in headers)
    
    if has_headers:
         print("[INFO] Detected Semi-Structured Resume. Parsing.")
         structured_data = parse_structured_resume(english_text)
         if structured_data["name"] or structured_data["education"]:
             final_data = {
                "full_name": structured_data["name"] or "Your Name",
                "phone": structured_data["phone"],
                "email": structured_data["email"],
                "location": structured_data["location"] or "",
                "summary": structured_data["summary"],
                "education": "\n".join(structured_data["education"]), 
                "responsibilities": "\n".join(structured_data["responsibilities"]),
                "projects": "\n".join(structured_data["projects"]),
                "skills": ", ".join(structured_data["skills"]),
                "certifications": "\n".join(structured_data["certifications"]),
            }
             return {"resume": build_resume_string(final_data)}
    
    # If not structured, or parsing failed, Try AI
    ai_data = generate_content_ai(raw_text, "resume")
    
    if ai_data:
        # Convert AI Structure (New Enriched Format) -> User Structure
        user_data = ai_data.copy()
        
        # Map new keys to internal keys if present
        if "professional_summary" in user_data:
            user_data["summary"] = user_data["professional_summary"]
            
        if "roles_and_responsibilities" in user_data:
            user_data["responsibilities"] = user_data["roles_and_responsibilities"]
            
        # Merge technical and tools into one 'skills' list for the template if separate
        tech_skills = user_data.get("technical_skills", [])
        tools = user_data.get("tools_and_technologies", [])
        soft = user_data.get("soft_skills", [])
        
        # Ensure they are lists
        if isinstance(tech_skills, str): tech_skills = [tech_skills]
        if isinstance(tools, str): tools = [tools]
        
        # Combine for the main 'Skills' section
        all_skills = []
        if tech_skills: all_skills.extend(tech_skills if isinstance(tech_skills, list) else [])
        if tools: all_skills.extend(tools if isinstance(tools, list) else [])
        
        user_data["skills"] = all_skills
        
        final_data = build_resume_data(user_data)
        return {"resume": build_resume_string(final_data)}

    # Fallback Parsing (Unstructured Regex)
    print("[WARN] AI Failed or Key Missing. Using Regex Fallback.")
    parsed = extract_fields(english_text)
    
    # Generic Fallback if Regex also fails badly
    # Use parse_structured_resume one last time just in case loose headers exist
    fallback_struct = parse_structured_resume(english_text)
    if fallback_struct["name"] or fallback_struct["skills"]:
         final_data = {
            "full_name": fallback_struct["name"] or parsed["name"],
            "phone": fallback_struct["phone"] or parsed["phone"],
            "email": fallback_struct["email"] or parsed["mailid"],
            "location": fallback_struct["location"] or "",
            "summary": fallback_struct["summary"],
            "education": "\n".join(fallback_struct["education"]),
            "responsibilities": "\n".join(fallback_struct["responsibilities"]),
            "projects": "\n".join(fallback_struct["projects"]),
            "skills": ", ".join(fallback_struct["skills"]),
        }
         return {"resume": build_resume_string(final_data)}

    # Last Resort
    user_data = {
        "name": parsed["name"],
        "email": parsed["mailid"],
        "phone": parsed["phone"],
        "location": "India", 
        "education": [], 
        "skills": [s.strip() for s in parsed.get("skills", "").split(",")],
    }
    final_data = build_resume_data(user_data)
    # USE NEW BUILD STRING FUNCTION
    # USE NEW BUILD STRING FUNCTION
    return {"resume": build_resume_string(final_data)}


# ================================
# PDF TOOLKIT ENDPOINTS
# ================================
from fastapi import UploadFile, File
import shutil
import tempfile
from PyPDF2 import PdfMerger, PdfReader, PdfWriter

@app.post("/pdf/merge")
async def merge_pdfs(files: List[UploadFile] = File(...)):
    merger = PdfMerger()
    temp_dir = tempfile.mkdtemp()
    try:
        paths = []
        for file in files:
            file_path = os.path.join(temp_dir, file.filename)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            merger.append(file_path)
        
        output_path = os.path.join(temp_dir, "merged.pdf")
        merger.write(output_path)
        merger.close()
        
        return StreamingResponse(
            open(output_path, "rb"),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=merged.pdf"}
        )
    except Exception as e:
        return {"error": str(e)}

@app.post("/pdf/split")
async def split_pdf(file: UploadFile = File(...)):
    reader = PdfReader(file.file)
    temp_dir = tempfile.mkdtemp()
    try:
        zip_path = os.path.join(temp_dir, "split_pages.zip")
        import zipfile
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for i in range(len(reader.pages)):
                writer = PdfWriter()
                writer.add_page(reader.pages[i])
                page_path = os.path.join(temp_dir, f"page_{i+1}.pdf")
                with open(page_path, "wb") as f:
                    writer.write(f)
                zipf.write(page_path, f"page_{i+1}.pdf")
        
        return StreamingResponse(
            open(zip_path, "rb"),
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=split_pages.zip"}
        )
    except Exception as e:
        import traceback
        print(f"Split PDF Error: {traceback.format_exc()}")
        return {"error": str(e)}

@app.post("/pdf/to-word")
async def pdf_to_word(file: UploadFile = File(...)):
    """PDF to Word with Engine-level Exact Match."""
    temp_dir = tempfile.mkdtemp(dir=BASE_CONVERSION_DIR)
    try:
        pdf_path = os.path.join(temp_dir, file.filename)
        with open(pdf_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # 1. Try LibreOffice Engine (Highest Quality)
        output_path = convert_with_libreoffice(pdf_path, "docx")
        
        if not output_path:
            # 2. Fallback to Deep Match internal logic
            print("[INFO] Engine failed. Using internal pdf2docx fallback.")
            docx_path = os.path.join(temp_dir, "output.docx")
            from pdf2docx import Converter
            cv = Converter(pdf_path)
            cv.convert(docx_path, start=0, end=None) 
            cv.close()
            output_path = docx_path
        
        return StreamingResponse(
            open(output_path, "rb"),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename={os.path.basename(output_path)}"}
        )
    except Exception as e:
        import traceback
        print(f"PDF to Word Error: {traceback.format_exc()}")
        return {"error": f"Failed to convert PDF to Word: {str(e)}"}

from docx import Document as DocxDocument

def _convert_docx_to_pdf_internal(docx_path, pdf_path):
    """
    Deep Match Conversion: Extracts exact font sizes, colors, and layout.
    """
    from docx.shared import Pt, RGBColor
    from docx.enum.text import WD_ALIGN_PARAGRAPH
    doc = DocxDocument(docx_path)
    styles = _get_styles()
    elements = []
    
    # helper for hex conversion
    def rgb_to_hex(rgb_color):
        if not rgb_color: return colors.black
        try:
            return colors.HexColor(f"#{rgb_color}")
        except:
            return colors.black

    # document-wide defaults
    default_font_size = 10
    
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table as DocxTable
    from docx.text.paragraph import Paragraph as DocxPara

    for block in doc.element.body:
        if isinstance(block, CT_P):
            para = DocxPara(block, doc)
            if not para.text.strip():
                elements.append(Spacer(1, 6))
                continue
            
            # 1. Alignment & Spacing
            align = TA_LEFT
            if para.alignment == WD_ALIGN_PARAGRAPH.CENTER: align = TA_CENTER
            elif para.alignment == WD_ALIGN_PARAGRAPH.RIGHT: align = TA_RIGHT
            elif para.alignment == WD_ALIGN_PARAGRAPH.JUSTIFY: align = TA_JUSTIFY
            
            # Detect indents
            left_indent = para.paragraph_format.left_indent.pt if para.paragraph_format.left_indent else 0
            space_before = para.paragraph_format.space_before.pt if para.paragraph_format.space_before else 0
            space_after = para.paragraph_format.space_after.pt if para.paragraph_format.space_after else 4
            
            if space_before > 0: elements.append(Spacer(1, space_before))

            # 2. Build Runs with Mixed Formatting
            full_text = ""
            for run in para.runs:
                if not run.text: continue
                # Style flags
                t = run.text.replace("<", "&lt;").replace(">", "&gt;") # Escape for XML
                
                f_size = run.font.size.pt if run.font.size else default_font_size
                f_color = rgb_to_hex(run.font.color.rgb) if run.font.color and run.font.color.rgb else colors.black
                
                # Apply tags
                if run.bold: t = f"<b>{t}</b>"
                if run.underline: t = f"<u>{t}</u>"
                if run.italic: t = f"<i>{t}</i>"
                
                # Apply Size and Color
                t = f'<font size="{f_size}" color="{f_color}">{t}</font>'
                full_text += t
            
            if full_text:
                custom_style = ParagraphStyle(
                    'Dynamic', 
                    parent=styles['Normal'], 
                    alignment=align, 
                    leftIndent=left_indent,
                    leading=max(12, f_size * 1.2) # Dynamic leading based on font size
                )
                elements.append(Paragraph(full_text, custom_style))
                elements.append(Spacer(1, space_after))
                
        elif isinstance(block, CT_Tbl):
            table = DocxTable(block, doc)
            table_data = []
            for row in table.rows:
                row_data = []
                for cell in row.cells:
                    cell_text = ""
                    for p in cell.paragraphs:
                        # Recursive run processing for cells
                        for run in p.runs:
                            t = run.text.replace("<", "&lt;").replace(">", "&gt;")
                            f_size = run.font.size.pt if run.font.size else 9
                            if run.bold: t = f"<b>{t}</b>"
                            cell_text += f'<font size="{f_size}">{t}</font>'
                    row_data.append(Paragraph(cell_text.strip(), styles['Normal']))
                table_data.append(row_data)
            
            if table_data:
                available_width = 500
                col_widths = [available_width / len(table_data[0])] * len(table_data[0])
                t = Table(table_data, colWidths=col_widths)
                t.setStyle(TableStyle([
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('LEFTPADDING', (0,0), (-1,-1), 2),
                    ('RIGHTPADDING', (0,0), (-1,-1), 2),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 10))
        
    pdf_doc = SimpleDocTemplate(
        pdf_path, pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    pdf_doc.build(elements)

@app.post("/pdf/word-to-pdf")
async def word_to_pdf(file: UploadFile = File(...)):
    """Word to PDF with Engine-level Exact Match."""
    temp_dir = tempfile.mkdtemp(dir=BASE_CONVERSION_DIR)
    try:
        docx_path = os.path.join(temp_dir, file.filename)
        with open(docx_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # 1. Try LibreOffice Engine (100% Exact Matching)
        output_path = convert_with_libreoffice(docx_path, "pdf")
        
        if not output_path:
            # 2. Fallback to Deep Match internal logic
            print("[INFO] Engine failed. Using internal ReportLab fallback.")
            pdf_path = os.path.join(temp_dir, "output.pdf")
            _convert_docx_to_pdf_internal(docx_path, pdf_path)
            output_path = pdf_path
        
        return StreamingResponse(
            open(output_path, "rb"),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={os.path.basename(output_path)}"}
        )
    except Exception as e:
        import traceback
        print(f"Word to PDF Error: {traceback.format_exc()}")
        return {"error": f"Conversion failed: {str(e)}"}

@app.post("/pdf/compress")
async def compress_pdf(file: UploadFile = File(...)):
    temp_dir = tempfile.mkdtemp()
    try:
        input_path = os.path.join(temp_dir, "input.pdf")
        output_path = os.path.join(temp_dir, "compressed.pdf")
        with open(input_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        reader = PdfReader(input_path)
        writer = PdfWriter()
        
        for page in reader.pages:
            page.compress_content_streams() # Basic compression
            writer.add_page(page)
            
        with open(output_path, "wb") as f:
            writer.write(f)
            
        return StreamingResponse(
            open(output_path, "rb"),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=compressed.pdf"}
        )
    except Exception as e:
        return {"error": str(e)}

@app.post("/pdf/to-excel")
async def pdf_to_excel(file: UploadFile = File(...)):
    """Improved PDF to Excel using structured extraction."""
    temp_dir = tempfile.mkdtemp()
    try:
        input_path = os.path.join(temp_dir, "input.pdf")
        output_path = os.path.join(temp_dir, "output.xlsx")
        with open(input_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        import pandas as pd
        from PyPDF2 import PdfReader
        
        reader = PdfReader(input_path)
        all_data = []
        
        # We attempt to find rows based on line spacing
        for page in reader.pages:
            text = page.extract_text()
            if not text: continue
            
            lines = text.split('\n')
            for line in lines:
                # Try to detect columns: commas, tabs, or large spaces
                if '\t' in line:
                    parts = line.split('\t')
                elif '    ' in line:
                    parts = re.split(r'\s{3,}', line)
                else:
                    parts = [line.strip()]
                
                if parts and any(p.strip() for p in parts):
                    all_data.append(parts)
        
        if not all_data:
            return {"error": "No text could be extracted from this PDF."}
            
        df = pd.DataFrame(all_data)
        df.to_excel(output_path, index=False, header=False)
        
        return StreamingResponse(
            open(output_path, "rb"),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=output.xlsx"}
        )
    except Exception as e:
        import traceback
        print(f"PDF to Excel Error: {traceback.format_exc()}")
        return {"error": str(e)}

@app.post("/pdf/excel-to-pdf")
async def excel_to_pdf(file: UploadFile = File(...)):
    """Professional Excel to PDF using Platypus Tables."""
    temp_dir = tempfile.mkdtemp()
    try:
        input_path = os.path.join(temp_dir, "input.xlsx")
        output_path = os.path.join(temp_dir, "output.pdf")
        with open(input_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        import pandas as pd
        df = pd.read_excel(input_path)
        df = df.fillna("") # Clean NaNs
        
        styles = _get_styles()
        elements = []
        
        # Add Header
        elements.append(Paragraph(f"Spreadsheet Export: {file.filename}", styles['HeaderName']))
        elements.append(Spacer(1, 20))
        
        # Prepare Table Data
        # Include headers as first row
        data = [df.columns.tolist()] + df.values.tolist()
        
        # Create Platypus Table
        # Auto-calculate width (max 500)
        col_count = len(df.columns)
        col_widths = [500 / col_count for _ in range(col_count)]
        
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2C3E50")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8 if col_count > 6 else 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        
        elements.append(t)
        
        return _create_platypus_pdf(elements, 'output.pdf')
        
    except Exception as e:
        import traceback
        print(f"Excel to PDF Error: {traceback.format_exc()}")
        return {"error": str(e)}

@app.post("/pdf/ppt-to-pdf")
async def ppt_to_pdf(file: UploadFile = File(...)):
    # Fallback implementation
    return {"error": "PowerPoint to PDF requires LibreOffice for reliable conversion. Basic text extraction coming soon!"}






@app.post("/generate-targeted-resume")
async def generate_targeted_resume(req: TargetedResumeRequest):
    """
    Generates a resume tailored to a specific job description.
    """
    try:
        # Prompt for targeted resume
        prompt = (
            "You are an expert Resume Strategist. Your task is to craft a HIGH-IMPACT, TARGETED RESUME "
            "by aligning the user's personal info and experience with a specific Job Description (JD).\n\n"
            "STRICT RULES:\n"
            "1. ALIGNMENT: Highlight skills and experiences from the personal info that are most relevant to the JD.\n"
            "2. KEYWORDS: Incorporate relevant keywords from the JD naturally into the resume.\n"
            "3. STRUCTURE: Use the professional resume format with Name, Contact, Summary, Education, Experience, Projects, and Skills.\n"
            "4. QUALITY: Use powerful action verbs and quantified achievements where possible.\n\n"
            "User's Personal Info/Experience:\n"
            f"<<< {req.personal_info} >>>\n\n"
            "Target Job Description:\n"
            f"<<< {req.job_description} >>>\n\n"
            "Return the tailored resume as PLAIN TEXT in a professional layout. "
            "Start with the Name at the very top."
        )

        # Use Gemini or OpenAI/Mistral
        ai_data = None
        
        # Try Mistral/OpenAI first
        mistral_api_key = os.environ.get("MISTRAL_API_KEY", "ZVzUQWNuusFD28EPJZia3E6ecLrgC1Em")
        if mistral_api_key:
            try:
                from mistralai import Mistral
                client = Mistral(api_key=mistral_api_key)
                chat_response = client.chat.complete(
                    model="mistral-large-latest",
                    messages=[{"role": "user", "content": prompt}]
                )
                ai_data = chat_response.choices[0].message.content
            except: pass

        if not ai_data and openai.api_key:
            try:
                client = openai.OpenAI()
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role": "user", "content": prompt}]
                )
                ai_data = response.choices[0].message.content
            except: pass

        if not ai_data and gemini_api_key:
            try:
                model = genai.GenerativeModel('gemini-1.5-flash')
                response = model.generate_content(prompt)
                ai_data = response.text
            except: pass

        if not ai_data:
            return {"error": "AI generation failed. Please check your API keys."}

        return {"resume": ai_data.strip()}

    except Exception as e:
        print(f"Targeted Resume Error: {e}")
        return {"error": str(e)}
